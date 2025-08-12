import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import yfinance as yf
from datetime import datetime, timedelta
import sqlite3
from arch import arch_model
from scipy import stats
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference
import warnings
warnings.filterwarnings('ignore')

class PortfolioRiskManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Portfolio Risk Management System")
        self.root.geometry("1400x900")
        
        # Initialize data storage
        self.portfolio_data = {}
        self.risk_metrics = {}
        self.var_results = {}
        
        # Initialize database connection
        self.conn = sqlite3.connect('portfolio_data.db')
        
        # Create database
        self.init_database()
        
        # Create GUI
        self.create_gui()
        
    def init_database(self):
        """Initialize SQLite database for storing portfolio data"""
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stock_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                date DATE NOT NULL,
                open_price REAL,
                high_price REAL,
                low_price REAL,
                close_price REAL,
                volume INTEGER,
                adj_close REAL,
                UNIQUE(symbol, date)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS portfolio_weights (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                weight REAL NOT NULL,
                date_added DATE DEFAULT CURRENT_DATE
            )
        ''')
        
        self.conn.commit()
    
    def get_adj_close_column(self, data):
        """Extract adjusted close column from yfinance data, handling different formats"""
        try:
            # Handle different column naming conventions
            possible_columns = ['Adj Close', 'adj_close', 'adjclose', 'AdjClose']
            
            for col in possible_columns:
                if col in data.columns:
                    return data[col]
            
            # If no adjusted close found, use regular close
            if 'Close' in data.columns:
                print("Warning: Using 'Close' instead of 'Adj Close'")
                return data['Close']
            elif 'close' in data.columns:
                print("Warning: Using 'close' instead of 'Adj Close'")
                return data['close']
            
            # Handle multi-level columns (sometimes yfinance returns these)
            if hasattr(data.columns, 'levels'):
                for col in data.columns:
                    if isinstance(col, tuple) and len(col) > 1:
                        if 'adj close' in str(col).lower() or 'close' in str(col).lower():
                            return data[col]
            
            print(f"Available columns: {list(data.columns)}")
            return None
            
        except Exception as e:
            print(f"Error extracting adjusted close column: {str(e)}")
            return None
        
    def create_gui(self):
        """Create the main GUI interface"""
        # Create notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Tab 1: Data Input
        self.data_frame = ttk.Frame(notebook)
        notebook.add(self.data_frame, text="Data Input")
        self.create_data_input_tab()
        
        # Tab 2: Risk Analysis
        self.risk_frame = ttk.Frame(notebook)
        notebook.add(self.risk_frame, text="Risk Analysis")
        self.create_risk_analysis_tab()
        
        # Tab 3: Stress Testing
        self.stress_frame = ttk.Frame(notebook)
        notebook.add(self.stress_frame, text="Stress Testing")
        self.create_stress_testing_tab()
        
        # Tab 4: Reports
        self.reports_frame = ttk.Frame(notebook)
        notebook.add(self.reports_frame, text="Reports & Export")
        self.create_reports_tab()
        
    def create_data_input_tab(self):
        """Create data input interface"""
        # Portfolio composition frame
        portfolio_frame = ttk.LabelFrame(self.data_frame, text="Portfolio Composition", padding=10)
        portfolio_frame.pack(fill='x', padx=10, pady=5)
        
        # Stock symbols input
        ttk.Label(portfolio_frame, text="Stock Symbols (comma-separated):").pack(anchor='w')
        self.symbols_entry = ttk.Entry(portfolio_frame, width=50)
        self.symbols_entry.pack(fill='x', pady=2)
        self.symbols_entry.insert(0, "AAPL,GOOGL,MSFT,TSLA,AMZN")
        
        # Weights input
        ttk.Label(portfolio_frame, text="Weights (comma-separated, must sum to 1):").pack(anchor='w', pady=(10,0))
        self.weights_entry = ttk.Entry(portfolio_frame, width=50)
        self.weights_entry.pack(fill='x', pady=2)
        self.weights_entry.insert(0, "0.2,0.2,0.2,0.2,0.2")
        
        # Date range
        date_frame = ttk.Frame(portfolio_frame)
        date_frame.pack(fill='x', pady=10)
        
        ttk.Label(date_frame, text="Start Date:").pack(side='left')
        self.start_date = ttk.Entry(date_frame, width=12)
        self.start_date.pack(side='left', padx=5)
        self.start_date.insert(0, "2020-01-01")
        
        ttk.Label(date_frame, text="End Date:").pack(side='left', padx=(20,0))
        self.end_date = ttk.Entry(date_frame, width=12)
        self.end_date.pack(side='left', padx=5)
        self.end_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Load data button
        ttk.Button(portfolio_frame, text="Load Portfolio Data", 
                  command=self.load_portfolio_data).pack(pady=10)
        
        # Data display frame
        display_frame = ttk.LabelFrame(self.data_frame, text="Data Preview", padding=10)
        display_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Treeview for data display
        columns = ('Symbol', 'Current Price', 'Weight', '1D Return', '30D Volatility')
        self.data_tree = ttk.Treeview(display_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            self.data_tree.heading(col, text=col)
            self.data_tree.column(col, width=120)
            
        scrollbar = ttk.Scrollbar(display_frame, orient='vertical', command=self.data_tree.yview)
        self.data_tree.configure(yscrollcommand=scrollbar.set)
        
        self.data_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
    def create_risk_analysis_tab(self):
        """Create risk analysis interface"""
        # Controls frame
        controls_frame = ttk.LabelFrame(self.risk_frame, text="Risk Analysis Controls", padding=10)
        controls_frame.pack(fill='x', padx=10, pady=5)
        
        # VaR confidence levels
        ttk.Label(controls_frame, text="VaR Confidence Levels:").pack(anchor='w')
        confidence_frame = ttk.Frame(controls_frame)
        confidence_frame.pack(fill='x', pady=5)
        
        self.conf_99 = tk.BooleanVar(value=True)
        self.conf_95 = tk.BooleanVar(value=True)
        self.conf_90 = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(confidence_frame, text="99%", variable=self.conf_99).pack(side='left')
        ttk.Checkbutton(confidence_frame, text="95%", variable=self.conf_95).pack(side='left', padx=20)
        ttk.Checkbutton(confidence_frame, text="90%", variable=self.conf_90).pack(side='left')
        
        # Analysis buttons
        button_frame = ttk.Frame(controls_frame)
        button_frame.pack(fill='x', pady=10)
        
        ttk.Button(button_frame, text="Calculate VaR (Historical)", 
                  command=self.calculate_historical_var).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Calculate VaR (Monte Carlo)", 
                  command=self.calculate_monte_carlo_var).pack(side='left', padx=5)
        if arch_model:
            ttk.Button(button_frame, text="GARCH Volatility Model", 
                      command=self.run_garch_model).pack(side='left', padx=5)
        
        # Results display frame
        results_frame = ttk.LabelFrame(self.risk_frame, text="Risk Metrics", padding=10)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Create matplotlib figure
        self.risk_fig, ((self.ax1, self.ax2), (self.ax3, self.ax4)) = plt.subplots(2, 2, figsize=(12, 8))
        self.risk_canvas = FigureCanvasTkAgg(self.risk_fig, results_frame)
        self.risk_canvas.get_tk_widget().pack(fill='both', expand=True)
        
    def create_stress_testing_tab(self):
        """Create stress testing interface"""
        # Stress test controls
        controls_frame = ttk.LabelFrame(self.stress_frame, text="Stress Test Scenarios", padding=10)
        controls_frame.pack(fill='x', padx=10, pady=5)
        
        # Historical scenarios
        hist_frame = ttk.LabelFrame(controls_frame, text="Historical Scenarios", padding=5)
        hist_frame.pack(fill='x', pady=5)
        
        scenarios = [
            ("2008 Financial Crisis", "2008-09-01", "2009-03-01"),
            ("COVID-19 Pandemic", "2020-02-01", "2020-04-01"),
            ("Dot-com Bubble", "2000-03-01", "2002-10-01")
        ]
        
        self.scenario_vars = {}
        for name, start, end in scenarios:
            var = tk.BooleanVar()
            self.scenario_vars[name] = {'var': var, 'start': start, 'end': end}
            ttk.Checkbutton(hist_frame, text=f"{name} ({start} to {end})", 
                           variable=var).pack(anchor='w')
        
        # Hypothetical shocks
        shock_frame = ttk.LabelFrame(controls_frame, text="Hypothetical Shocks", padding=5)
        shock_frame.pack(fill='x', pady=5)
        
        ttk.Label(shock_frame, text="Market Shock (%):").pack(anchor='w')
        self.market_shock = ttk.Entry(shock_frame, width=10)
        self.market_shock.pack(anchor='w', pady=2)
        self.market_shock.insert(0, "-20")
        
        ttk.Label(shock_frame, text="Interest Rate Change (bps):").pack(anchor='w', pady=(10,0))
        self.rate_shock = ttk.Entry(shock_frame, width=10)
        self.rate_shock.pack(anchor='w', pady=2)
        self.rate_shock.insert(0, "200")
        
        ttk.Button(controls_frame, text="Run Stress Tests", 
                  command=self.run_stress_tests).pack(pady=10)
        
        # Stress test results
        stress_results_frame = ttk.LabelFrame(self.stress_frame, text="Stress Test Results", padding=10)
        stress_results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.stress_fig, (self.stress_ax1, self.stress_ax2) = plt.subplots(1, 2, figsize=(12, 6))
        self.stress_canvas = FigureCanvasTkAgg(self.stress_fig, stress_results_frame)
        self.stress_canvas.get_tk_widget().pack(fill='both', expand=True)
        
    def create_reports_tab(self):
        """Create reports and export interface"""
        # Export controls
        export_frame = ttk.LabelFrame(self.reports_frame, text="Export Options", padding=10)
        export_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(export_frame, text="Generate Excel Report", 
                  command=self.generate_excel_report).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export Charts", 
                  command=self.export_charts).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Save Risk Dashboard", 
                  command=self.save_dashboard).pack(side='left', padx=5)
        
        # Summary statistics
        summary_frame = ttk.LabelFrame(self.reports_frame, text="Portfolio Summary", padding=10)
        summary_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.summary_text = tk.Text(summary_frame, height=20, width=80)
        summary_scrollbar = ttk.Scrollbar(summary_frame, orient='vertical', command=self.summary_text.yview)
        self.summary_text.configure(yscrollcommand=summary_scrollbar.set)
        
        self.summary_text.pack(side='left', fill='both', expand=True)
        summary_scrollbar.pack(side='right', fill='y')
        
    def load_portfolio_data(self):
        """Load portfolio data from Yahoo Finance"""
        try:
            symbols = [s.strip().upper() for s in self.symbols_entry.get().split(',')]
            weights = [float(w.strip()) for w in self.weights_entry.get().split(',')]
            
            if len(symbols) != len(weights):
                messagebox.showerror("Error", "Number of symbols must match number of weights")
                return
                
            if abs(sum(weights) - 1.0) > 0.001:
                messagebox.showerror("Error", "Weights must sum to 1.0")
                return
            
            start_date = self.start_date.get()
            end_date = self.end_date.get()
            
            # Download data with better error handling
            self.portfolio_data = {}
            failed_symbols = []
            
            for symbol in symbols:
                try:
                    print(f"Downloading data for {symbol}...")
                    data = yf.download(symbol, start=start_date, end=end_date, progress=False)
                    
                    if data.empty:
                        failed_symbols.append(symbol)
                        continue
                    
                    # Handle multi-level columns
                    if hasattr(data.columns, 'levels'):
                        data.columns = data.columns.droplevel(1)
                    
                    # Verify we have the required data
                    adj_close = self.get_adj_close_column(data)
                    if adj_close is None:
                        failed_symbols.append(symbol)
                        continue
                    
                    # Store the cleaned data
                    self.portfolio_data[symbol] = data
                    
                    # Store in database
                    self.store_price_data(symbol, data)
                    
                except Exception as e:
                    print(f"Failed to download {symbol}: {str(e)}")
                    failed_symbols.append(symbol)
                    continue
            
            if failed_symbols:
                messagebox.showwarning("Warning", f"Failed to load data for: {', '.join(failed_symbols)}")
            
            if not self.portfolio_data:
                messagebox.showerror("Error", "No data could be loaded for any symbols")
                return
            
            # Update weights to only include successful symbols
            successful_symbols = list(self.portfolio_data.keys())
            original_weights = dict(zip(symbols, weights))
            
            # Recalculate weights for successful symbols only
            total_weight = sum(original_weights[symbol] for symbol in successful_symbols)
            self.weights = {symbol: original_weights[symbol] / total_weight for symbol in successful_symbols}
            
            # Update display
            self.update_data_display()
            
            messagebox.showinfo("Success", f"Portfolio data loaded successfully for {len(successful_symbols)} symbols!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
    
    def store_price_data(self, symbol, data):
        """Store price data in database"""
        cursor = self.conn.cursor()
        
        adj_close = self.get_adj_close_column(data)
        if adj_close is None:
            return
        
        for date, row in data.iterrows():
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO stock_prices 
                    (symbol, date, open_price, high_price, low_price, close_price, volume, adj_close)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (symbol, date.strftime('%Y-%m-%d'), 
                      row.get('Open', 0), row.get('High', 0), 
                      row.get('Low', 0), row.get('Close', 0), 
                      row.get('Volume', 0), adj_close.loc[date] if date in adj_close.index else 0))
            except Exception as e:
                print(f"Error storing data for {symbol} on {date}: {str(e)}")
                continue
        
        self.conn.commit()
    
    def update_data_display(self):
        """Update the data display treeview"""
        # Clear existing data
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)
        
        for symbol, data in self.portfolio_data.items():
            adj_close = self.get_adj_close_column(data)
            if adj_close is None:
                continue
                
            current_price = adj_close.iloc[-1]
            daily_return = adj_close.pct_change().iloc[-1] * 100
            volatility = adj_close.pct_change().rolling(30).std().iloc[-1] * np.sqrt(252) * 100
            weight = self.weights[symbol] * 100
            
            self.data_tree.insert('', 'end', values=(
                symbol, 
                f"${current_price:.2f}", 
                f"{weight:.1f}%", 
                f"{daily_return:.2f}%" if not np.isnan(daily_return) else "N/A", 
                f"{volatility:.2f}%" if not np.isnan(volatility) else "N/A"
            ))
    
    def calculate_historical_var(self):
        """Calculate Value at Risk using historical simulation"""
        if not self.portfolio_data:
            messagebox.showerror("Error", "Please load portfolio data first")
            return
        
        try:
            # Calculate portfolio returns
            returns_data = {}
            for symbol, data in self.portfolio_data.items():
                adj_close = self.get_adj_close_column(data)
                if adj_close is not None:
                    returns_data[symbol] = adj_close.pct_change().dropna()
            
            if not returns_data:
                messagebox.showerror("Error", "No valid return data available")
                return
            
            returns_df = pd.DataFrame(returns_data)
            
            # Calculate portfolio returns
            weights_array = np.array([self.weights[symbol] for symbol in returns_df.columns])
            portfolio_returns = (returns_df * weights_array).sum(axis=1)
            
            # Calculate VaR at different confidence levels
            var_results = {}
            confidence_levels = []
            
            if self.conf_99.get():
                confidence_levels.append(0.01)
            if self.conf_95.get():
                confidence_levels.append(0.05)
            if self.conf_90.get():
                confidence_levels.append(0.10)
            
            for alpha in confidence_levels:
                var_value = -np.percentile(portfolio_returns, alpha * 100)
                es_value = -portfolio_returns[portfolio_returns <= -var_value].mean()
                var_results[f"{(1-alpha)*100:.0f}%"] = {
                    'VaR': var_value,
                    'Expected Shortfall': es_value
                }
            
            self.var_results = var_results
            
            # Plot results
            self.plot_var_results(portfolio_returns, var_results)
            
            # Update summary
            self.update_summary("Historical VaR", var_results)
            
        except Exception as e:
            messagebox.showerror("Error", f"VaR calculation failed: {str(e)}")
    
    def plot_var_results(self, portfolio_returns, var_results):
        """Plot VaR analysis results"""
        self.ax1.clear()
        self.ax2.clear()
        self.ax3.clear()
        self.ax4.clear()
        
        # Plot 1: Return distribution
        self.ax1.hist(portfolio_returns, bins=50, alpha=0.7, density=True)
        for conf, metrics in var_results.items():
            self.ax1.axvline(-metrics['VaR'], color='red', linestyle='--', 
                           label=f"VaR {conf}: {metrics['VaR']:.4f}")
        self.ax1.set_title('Portfolio Return Distribution')
        self.ax1.set_xlabel('Daily Returns')
        self.ax1.set_ylabel('Density')
        self.ax1.legend()
        
        # Plot 2: Cumulative returns
        cumulative_returns = (1 + portfolio_returns).cumprod()
        self.ax2.plot(cumulative_returns.index, cumulative_returns.values)
        self.ax2.set_title('Cumulative Portfolio Returns')
        self.ax2.set_xlabel('Date')
        self.ax2.set_ylabel('Cumulative Return')
        
        # Plot 3: Rolling volatility
        rolling_vol = portfolio_returns.rolling(30).std() * np.sqrt(252)
        self.ax3.plot(rolling_vol.index, rolling_vol.values)
        self.ax3.set_title('30-Day Rolling Volatility')
        self.ax3.set_xlabel('Date')
        self.ax3.set_ylabel('Annualized Volatility')
        
        # Plot 4: Correlation heatmap
        returns_data = {}
        for symbol, data in self.portfolio_data.items():
            adj_close = self.get_adj_close_column(data)
            if adj_close is not None:
                returns_data[symbol] = adj_close.pct_change().dropna()
        
        if returns_data:
            returns_df = pd.DataFrame(returns_data)
            correlation_matrix = returns_df.corr()
            
            im = self.ax4.imshow(correlation_matrix, cmap='coolwarm', aspect='auto', vmin=-1, vmax=1)
            self.ax4.set_xticks(range(len(correlation_matrix.columns)))
            self.ax4.set_yticks(range(len(correlation_matrix.columns)))
            self.ax4.set_xticklabels(correlation_matrix.columns, rotation=45)
            self.ax4.set_yticklabels(correlation_matrix.columns)
            self.ax4.set_title('Asset Correlation Matrix')
            
            # Add correlation values
            for i in range(len(correlation_matrix.columns)):
                for j in range(len(correlation_matrix.columns)):
                    self.ax4.text(j, i, f'{correlation_matrix.iloc[i, j]:.2f}', 
                                ha='center', va='center', fontsize=8)
        
        plt.tight_layout()
        self.risk_canvas.draw()
    
    def calculate_monte_carlo_var(self):
        """Calculate VaR using Monte Carlo simulation"""
        if not self.portfolio_data:
            messagebox.showerror("Error", "Please load portfolio data first")
            return
        
        try:
            # Calculate historical statistics
            returns_data = {}
            for symbol, data in self.portfolio_data.items():
                adj_close = self.get_adj_close_column(data)
                if adj_close is not None:
                    returns_data[symbol] = adj_close.pct_change().dropna()
            
            if not returns_data:
                messagebox.showerror("Error", "No valid return data available")
                return
            
            returns_df = pd.DataFrame(returns_data)
            mean_returns = returns_df.mean()
            cov_matrix = returns_df.cov()
            
            # Monte Carlo simulation
            num_simulations = 10000
            weights_array = np.array([self.weights[symbol] for symbol in returns_df.columns])
            
            # Generate random returns
            simulated_returns = np.random.multivariate_normal(
                mean_returns, cov_matrix, num_simulations
            )
            
            # Calculate portfolio returns
            portfolio_returns = np.dot(simulated_returns, weights_array)
            
            # Calculate VaR
            var_results = {}
            confidence_levels = []
            
            if self.conf_99.get():
                confidence_levels.append(0.01)
            if self.conf_95.get():
                confidence_levels.append(0.05)
            if self.conf_90.get():
                confidence_levels.append(0.10)
            
            for alpha in confidence_levels:
                var_value = -np.percentile(portfolio_returns, alpha * 100)
                es_value = -portfolio_returns[portfolio_returns <= -var_value].mean()
                var_results[f"{(1-alpha)*100:.0f}%"] = {
                    'VaR': var_value,
                    'Expected Shortfall': es_value
                }
            
            self.var_results = var_results
            
            # Update summary
            self.update_summary("Monte Carlo VaR", var_results)
            
            messagebox.showinfo("Success", "Monte Carlo VaR calculated successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Monte Carlo VaR calculation failed: {str(e)}")
    
    def run_garch_model(self):
        """Run GARCH model for volatility forecasting"""
        if not arch_model:
            messagebox.showerror("Error", "ARCH library not available. Please install with: pip install arch")
            return
            
        if not self.portfolio_data:
            messagebox.showerror("Error", "Please load portfolio data first")
            return
        
        try:
            # Calculate portfolio returns
            returns_data = {}
            for symbol, data in self.portfolio_data.items():
                adj_close = self.get_adj_close_column(data)
                if adj_close is not None:
                    returns_data[symbol] = adj_close.pct_change().dropna()
            
            if not returns_data:
                messagebox.showerror("Error", "No valid return data available")
                return
            
            returns_df = pd.DataFrame(returns_data)
            weights_array = np.array([self.weights[symbol] for symbol in returns_df.columns])
            portfolio_returns = (returns_df * weights_array).sum(axis=1)
            
            # Fit GARCH(1,1) model
            portfolio_returns_percent = portfolio_returns * 100  # Convert to percentage
            model = arch_model(portfolio_returns_percent, vol='Garch', p=1, q=1)
            fitted_model = model.fit(disp='off')
            
            # Forecast volatility
            forecast = fitted_model.forecast(horizon=30)
            forecasted_vol = np.sqrt(forecast.variance.iloc[-1, :]) / 100  # Convert back to decimal
            
            # Plot GARCH results
            self.plot_garch_results(portfolio_returns, fitted_model, forecasted_vol)
            
            messagebox.showinfo("Success", "GARCH model fitted successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"GARCH model failed: {str(e)}")
    
    def plot_garch_results(self, returns, fitted_model, forecasted_vol):
        """Plot GARCH model results"""
        self.ax1.clear()
        self.ax2.clear()
        
        # Plot conditional volatility
        conditional_vol = fitted_model.conditional_volatility / 100
        self.ax1.plot(returns.index, conditional_vol)
        self.ax1.set_title('GARCH Conditional Volatility')
        self.ax1.set_xlabel('Date')
        self.ax1.set_ylabel('Volatility')
        
        # Plot volatility forecast
        forecast_dates = pd.date_range(start=returns.index[-1], periods=31)[1:]
        self.ax2.plot(forecast_dates, forecasted_vol, 'r-', label='Forecasted Volatility')
        self.ax2.set_title('30-Day Volatility Forecast')
        self.ax2.set_xlabel('Date')
        self.ax2.set_ylabel('Volatility')
        self.ax2.legend()
        
        plt.tight_layout()
        self.risk_canvas.draw()
    
    def run_stress_tests(self):
        """Run stress tests on the portfolio"""
        if not self.portfolio_data:
            messagebox.showerror("Error", "Please load portfolio data first")
            return
        
        try:
            stress_results = {}
            
            # Historical scenario stress tests
            for scenario_name, scenario_data in self.scenario_vars.items():
                if scenario_data['var'].get():
                    start_date = scenario_data['start']
                    end_date = scenario_data['end']
                    
                    scenario_returns = self.calculate_scenario_impact(start_date, end_date)
                    stress_results[scenario_name] = scenario_returns
            
            # Hypothetical shock tests
            try:
                market_shock = float(self.market_shock.get()) / 100
                rate_shock = float(self.rate_shock.get()) / 10000  # Convert bps to decimal
                
                shock_impact = self.calculate_shock_impact(market_shock, rate_shock)
                stress_results['Market Shock'] = shock_impact
            except ValueError:
                messagebox.showwarning("Warning", "Invalid shock values entered")
            
            if not stress_results:
                messagebox.showwarning("Warning", "No stress test scenarios were run")
                return
            
            # Plot stress test results
            self.plot_stress_results(stress_results)
            
            messagebox.showinfo("Success", "Stress tests completed successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Stress test failed: {str(e)}")
    
    def calculate_scenario_impact(self, start_date, end_date):
        """Calculate portfolio impact during historical scenario"""
        scenario_returns = {}
        
        print(f"Calculating scenario impact from {start_date} to {end_date}")
        
        for symbol, data in self.portfolio_data.items():
            try:
                # Use get_adj_close_column instead of direct access
                adj_close = self.get_adj_close_column(data)
                if adj_close is None:
                    print(f"No price data available for {symbol}")
                    continue
                
                # Filter data for the scenario period
                scenario_data = adj_close.loc[start_date:end_date]
                
                if len(scenario_data) > 1:
                    start_price = scenario_data.iloc[0]
                    end_price = scenario_data.iloc[-1]
                    total_return = (end_price / start_price) - 1
                    scenario_returns[symbol] = total_return
                    print(f"{symbol}: {total_return:.4f} ({total_return*100:.2f}%)")
                else:
                    print(f"Insufficient data for {symbol} in period {start_date} to {end_date}")
                    
            except Exception as e:
                print(f"Error calculating scenario impact for {symbol}: {str(e)}")
                continue
        
        if not scenario_returns:
            print("No scenario returns calculated - check date ranges and data availability")
            return 0.0
        
        # Calculate weighted portfolio return
        portfolio_return = sum(scenario_returns[symbol] * self.weights[symbol] 
                             for symbol in scenario_returns.keys())
        
        print(f"Portfolio impact: {portfolio_return:.4f} ({portfolio_return*100:.2f}%)")
        return portfolio_return
    
    def calculate_shock_impact(self, market_shock, rate_shock):
        """Calculate impact of hypothetical shocks"""
        # Simplified shock calculation - apply market shock to all assets
        # In practice, you would use factor models and beta calculations
        portfolio_impact = market_shock  # Assuming beta = 1 for simplicity
        
        return portfolio_impact
    
    def plot_stress_results(self, stress_results):
        """Plot stress test results"""
        self.stress_ax1.clear()
        self.stress_ax2.clear()
        
        if not stress_results:
            self.stress_ax1.text(0.5, 0.5, 'No stress test results available\nCheck date ranges and data', 
                               ha='center', va='center', transform=self.stress_ax1.transAxes)
            self.stress_ax1.set_title('Stress Test Results - No Data')
            
            self.stress_ax2.text(0.5, 0.5, 'No VaR data available\nRun VaR analysis first', 
                               ha='center', va='center', transform=self.stress_ax2.transAxes)
            self.stress_ax2.set_title('VaR Analysis Required')
        else:
            # Bar chart of stress test results
            scenarios = list(stress_results.keys())
            impacts = [stress_results[scenario] * 100 for scenario in scenarios]  # Convert to percentage
            
            colors = ['red' if impact < 0 else 'green' for impact in impacts]
            bars = self.stress_ax1.bar(scenarios, impacts, color=colors, alpha=0.7)
            self.stress_ax1.set_title('Stress Test Results')
            self.stress_ax1.set_ylabel('Portfolio Impact (%)')
            self.stress_ax1.tick_params(axis='x', rotation=45)
            
            # Add value labels on bars
            for bar, impact in zip(bars, impacts):
                height = bar.get_height()
                self.stress_ax1.text(bar.get_x() + bar.get_width()/2., height,
                                   f'{impact:.2f}%', ha='center', va='bottom' if height >= 0 else 'top')
            
            # Risk decomposition pie chart
            if self.var_results:
                var_values = [metrics['VaR'] * 100 for metrics in self.var_results.values()]
                labels = list(self.var_results.keys())
                
                self.stress_ax2.pie(var_values, labels=labels, autopct='%1.1f%%')
                self.stress_ax2.set_title('VaR by Confidence Level')
            else:
                self.stress_ax2.text(0.5, 0.5, 'No VaR data available\nRun VaR analysis first', 
                                   ha='center', va='center', transform=self.stress_ax2.transAxes)
                self.stress_ax2.set_title('VaR Analysis Required')
        
        plt.tight_layout()
        self.stress_canvas.draw()
    
    def update_summary(self, analysis_type, results):
        """Update the summary text widget"""
        self.summary_text.insert(tk.END, f"\n=== {analysis_type} Results ===\n")
        
        for confidence, metrics in results.items():
            self.summary_text.insert(tk.END, f"\n{confidence} Confidence Level:\n")
            for metric, value in metrics.items():
                self.summary_text.insert(tk.END, f"  {metric}: {value:.4f} ({value*100:.2f}%)\n")
        
        self.summary_text.insert(tk.END, "\n" + "="*50 + "\n")
        self.summary_text.see(tk.END)
    
    def generate_excel_report(self):
        """Generate comprehensive Excel report with all calculation data and methodology"""
        if not self.portfolio_data:
            messagebox.showerror("Error", "Please load portfolio data first")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Excel Report"
            )
            
            if not filename:
                return
            
            wb = Workbook()
            
            # Portfolio Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Portfolio Summary"
            
            # Headers
            headers = ['Symbol', 'Weight', 'Current Price', 'Daily Return', '30D Volatility', 'Market Value', 'Price Column Used']
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Portfolio data with proper column detection
            total_value = 100000  # Assume $100k portfolio
            for row, (symbol, data) in enumerate(self.portfolio_data.items(), 2):
                # Use the helper method to get the correct price column
                adj_close_data = self.get_adj_close_column(data)
                price_column_name = 'Adj Close' if 'Adj Close' in data.columns else 'Close'
                
                if adj_close_data is not None:
                    current_price = adj_close_data.iloc[-1]
                    daily_return = adj_close_data.pct_change().iloc[-1]
                    volatility = adj_close_data.pct_change().rolling(30).std().iloc[-1] * np.sqrt(252)
                    market_value = total_value * self.weights[symbol]
                    
                    ws_summary.cell(row=row, column=1, value=symbol)
                    ws_summary.cell(row=row, column=2, value=self.weights[symbol])
                    ws_summary.cell(row=row, column=3, value=current_price)
                    ws_summary.cell(row=row, column=4, value=daily_return)
                    ws_summary.cell(row=row, column=5, value=volatility)
                    ws_summary.cell(row=row, column=6, value=market_value)
                    ws_summary.cell(row=row, column=7, value=price_column_name)
            
            # Calculation Methodology Sheet
            ws_methodology = wb.create_sheet("Calculation Methodology")
            methodology_data = [
                ["Calculation", "Formula", "Description"],
                ["Daily Return", "(Price_t - Price_t-1) / Price_t-1", "Percentage change in price from previous day"],
                ["30-Day Volatility", "std(daily_returns) * sqrt(252)", "Annualized volatility based on 30-day rolling window"],
                ["Historical VaR", "percentile(returns, confidence_level)", "Value at Risk using historical simulation"],
                ["Monte Carlo VaR", "percentile(simulated_returns, confidence_level)", "VaR using Monte Carlo simulation with 10,000 iterations"],
                ["Expected Shortfall", "mean(returns[returns <= VaR])", "Average of losses exceeding VaR threshold"],
                ["GARCH Volatility", "GARCH(1,1) model", "Conditional volatility using GARCH modeling"],
                ["Portfolio Return", "sum(weight_i * return_i)", "Weighted sum of individual asset returns"],
                ["Portfolio Volatility", "sqrt(w'Σw)", "Portfolio risk using covariance matrix"],
                ["Stress Test Impact", "(scenario_price - current_price) / current_price", "Percentage impact under stress scenarios"]
            ]
            
            for row, data_row in enumerate(methodology_data, 1):
                for col, value in enumerate(data_row, 1):
                    cell = ws_methodology.cell(row=row, column=col, value=value)
                    if row == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Risk Metrics Sheet with detailed calculations
            if self.var_results:
                ws_risk = wb.create_sheet("Risk Metrics")
                
                # Headers
                risk_headers = ["Confidence Level", "VaR", "VaR %", "Expected Shortfall", "ES %", "Calculation Method"]
                for col, header in enumerate(risk_headers, 1):
                    cell = ws_risk.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for row, (conf, metrics) in enumerate(self.var_results.items(), 2):
                    ws_risk.cell(row=row, column=1, value=conf)
                    ws_risk.cell(row=row, column=2, value=metrics['VaR'])
                    ws_risk.cell(row=row, column=3, value=metrics['VaR'] * 100)
                    ws_risk.cell(row=row, column=4, value=metrics['Expected Shortfall'])
                    ws_risk.cell(row=row, column=5, value=metrics['Expected Shortfall'] * 100)
                    ws_risk.cell(row=row, column=6, value="Historical Simulation")
            
            # Historical Data Sheet with all price and return data
            ws_data = wb.create_sheet("Historical Data")
            
            # Combine all price data using proper column detection
            all_data = pd.DataFrame()
            date_column_added = False
            
            for symbol, data in self.portfolio_data.items():
                adj_close_data = self.get_adj_close_column(data)
                if adj_close_data is not None:
                    if not date_column_added:
                        all_data['Date'] = adj_close_data.index
                        date_column_added = True
                    all_data[f"{symbol}_Price"] = adj_close_data.values
                    all_data[f"{symbol}_Return"] = adj_close_data.pct_change().values
                    all_data[f"{symbol}_30D_Vol"] = adj_close_data.pct_change().rolling(30).std().values * np.sqrt(252)
            
            # Write headers
            for col, column_name in enumerate(all_data.columns, 1):
                cell = ws_data.cell(row=1, column=col, value=column_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            for row, data_row in enumerate(all_data.itertuples(index=False), 2):
                for col, value in enumerate(data_row, 1):
                    if col == 1 and hasattr(value, 'strftime'):  # Date column
                        ws_data.cell(row=row, column=col, value=value.strftime('%Y-%m-%d'))
                    else:
                        ws_data.cell(row=row, column=col, value=value)
            
            # Portfolio Statistics Sheet
            ws_stats = wb.create_sheet("Portfolio Statistics")
            
            # Calculate portfolio-level statistics
            portfolio_returns = []
            for date in all_data['Date']:
                daily_return = 0
                for symbol in self.weights.keys():
                    if f"{symbol}_Return" in all_data.columns:
                        symbol_return = all_data[all_data['Date'] == date][f"{symbol}_Return"].iloc[0]
                        if not pd.isna(symbol_return):
                            daily_return += self.weights[symbol] * symbol_return
                portfolio_returns.append(daily_return)
            
            portfolio_stats = [
                ["Statistic", "Value", "Calculation"],
                ["Portfolio Mean Return", np.mean(portfolio_returns), "Average of daily portfolio returns"],
                ["Portfolio Volatility", np.std(portfolio_returns) * np.sqrt(252), "Annualized standard deviation"],
                ["Sharpe Ratio", np.mean(portfolio_returns) / np.std(portfolio_returns) * np.sqrt(252), "(Return - Risk-free) / Volatility"],
                ["Maximum Drawdown", np.min(np.cumsum(portfolio_returns)), "Largest peak-to-trough decline"],
                ["Skewness", stats.skew(portfolio_returns), "Measure of return distribution asymmetry"],
                ["Kurtosis", stats.kurtosis(portfolio_returns), "Measure of return distribution tail thickness"]
            ]
            
            for row, data_row in enumerate(portfolio_stats, 1):
                for col, value in enumerate(data_row, 1):
                    cell = ws_stats.cell(row=row, column=col, value=value)
                    if row == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Stress Test Results Sheet (if available)
            if hasattr(self, 'stress_results') and self.stress_results:
                ws_stress = wb.create_sheet("Stress Test Results")
                
                stress_headers = ["Scenario", "Impact %", "Portfolio Value", "Loss Amount", "Calculation Method"]
                for col, header in enumerate(stress_headers, 1):
                    cell = ws_stress.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for row, (scenario, impact) in enumerate(self.stress_results.items(), 2):
                    portfolio_value = 100000 * (1 + impact/100)
                    loss_amount = 100000 - portfolio_value
                    
                    ws_stress.cell(row=row, column=1, value=scenario)
                    ws_stress.cell(row=row, column=2, value=impact)
                    ws_stress.cell(row=row, column=3, value=portfolio_value)
                    ws_stress.cell(row=row, column=4, value=loss_amount)
                    ws_stress.cell(row=row, column=5, value="Historical scenario analysis")
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Comprehensive Excel report saved to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel report: {str(e)}")
            print(f"Excel export error details: {e}")  # Debug output

    def export_charts(self):
        """Export charts as images"""
        try:
            directory = filedialog.askdirectory(title="Select Directory to Save Charts")
            if not directory:
                return
            
            # Save risk analysis charts
            self.risk_fig.savefig(f"{directory}/risk_analysis.png", dpi=300, bbox_inches='tight')
            
            # Save stress test charts
            self.stress_fig.savefig(f"{directory}/stress_tests.png", dpi=300, bbox_inches='tight')
            
            messagebox.showinfo("Success", f"Charts exported to {directory}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export charts: {str(e)}")
    
    def save_dashboard(self):
        """Save current dashboard state"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                title="Save Dashboard"
            )
            
            if not filename:
                return
            
            dashboard_data = {
                'symbols': list(self.weights.keys()),
                'weights': list(self.weights.values()),
                'var_results': self.var_results,
                'timestamp': datetime.now().isoformat()
            }
            
            import json
            with open(filename, 'w') as f:
                json.dump(dashboard_data, f, indent=2)
            
            messagebox.showinfo("Success", f"Dashboard saved to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save dashboard: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PortfolioRiskManager(root)
    root.mainloop()